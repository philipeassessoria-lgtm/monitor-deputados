"""
scraper_estaduais.py
Scraper generico para deputados estaduais de todas as 27 UFs.

Uso:
  python scraper_estaduais.py                 # todas UFs
  python scraper_estaduais.py --uf MG         # so MG
  python scraper_estaduais.py --uf SP,RJ,BA   # lista

Saida: deputados_estaduais.json + deputados_estaduais.xlsx

Dependencias:
  pip install requests beautifulsoup4 lxml openpyxl
"""

import argparse
import json
import sys
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup

sys.path.append(str(Path(__file__).parent))
from config_ufs import UFS_CONFIG

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Accept-Language": "pt-BR,pt;q=0.9",
}


def fetch_html(url: str, timeout=30):
    r = requests.get(url, headers=HEADERS, timeout=timeout)
    r.raise_for_status()
    return BeautifulSoup(r.text, "lxml")


def scrape_api_json(cfg):
    r = requests.get(cfg["url"], headers=HEADERS, timeout=30)
    r.raise_for_status()
    data = r.json()
    for chave in cfg.get("api_path", []):
        data = data.get(chave, [])
    mapa = cfg.get("mapa_campos", {})
    out = []
    for item in data:
        out.append({
            "nome": item.get(mapa.get("nome", "nome"), ""),
            "partido": item.get(mapa.get("partido", "partido"), ""),
            "email": item.get(mapa.get("email", "email"), ""),
            "instagram": "",
        })
    return out


def scrape_html(cfg):
    soup = fetch_html(cfg["url"])
    out = []
    items = soup.select(cfg.get("seletor_lista", ""))
    if not items:
        print(f"    [AVISO] Nenhum item encontrado com seletor '{cfg.get('seletor_lista')}'.")
        print(f"            Inspecione manualmente: {cfg['url']}")
        return out
    for it in items:
        nome_el = it.select_one(cfg.get("seletor_nome", ""))
        partido_el = it.select_one(cfg.get("seletor_partido", ""))
        nome = nome_el.get_text(strip=True) if nome_el else ""
        partido = partido_el.get_text(strip=True) if partido_el else ""
        if not nome:
            continue
        # Extrair possiveis links de Instagram ja na pagina
        ig = ""
        for a in it.select("a[href*='instagram.com']"):
            ig = a.get("href", "")
            break
        out.append({
            "nome": nome,
            "partido": partido,
            "email": "",
            "instagram": ig,
        })
    return out


def scrape_uf(uf: str):
    cfg = UFS_CONFIG.get(uf)
    if not cfg:
        return []
    print(f"\n[{uf}] {cfg['assembleia']} ({cfg['status']})")
    try:
        if cfg["tipo"] == "api_json":
            deps = scrape_api_json(cfg)
        elif cfg["tipo"] == "html":
            deps = scrape_html(cfg)
        else:
            print(f"    Tipo desconhecido: {cfg['tipo']}")
            return []
        print(f"    OK: {len(deps)} deputado(s) extraido(s)")
        return [{**d, "uf": uf, "assembleia": cfg["assembleia"]} for d in deps]
    except Exception as e:
        print(f"    ERRO: {e}")
        return []


def salvar_json(deps, path="deputados_estaduais.json"):
    Path(path).write_text(json.dumps(deps, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nSalvo em {path}: {len(deps)} registros")


def salvar_xlsx(deps, path="deputados_estaduais.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Deputados Estaduais"
    cols = ["UF", "Assembleia", "Nome", "Partido", "Email", "Instagram"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for d in deps:
        ws.append([d["uf"], d["assembleia"], d["nome"], d["partido"],
                   d["email"], d["instagram"]])
    widths = [6, 12, 35, 10, 35, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Salvo em {path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--uf", help="UFs separadas por virgula (ex: MG,SP). Default: todas")
    ap.add_argument("--delay", type=float, default=2.0, help="Delay entre UFs (s)")
    args = ap.parse_args()

    ufs = args.uf.upper().split(",") if args.uf else list(UFS_CONFIG.keys())
    todos = []
    for uf in ufs:
        todos.extend(scrape_uf(uf))
        time.sleep(args.delay)

    salvar_json(todos)
    salvar_xlsx(todos)


if __name__ == "__main__":
    main()
