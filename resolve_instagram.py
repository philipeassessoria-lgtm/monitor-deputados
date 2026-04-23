"""
resolve_instagram.py
Descobre handle do Instagram de cada deputado usando busca DuckDuckGo HTML
(sem API key, sem custo). Refinamento manual sempre possivel depois.

Estrategia:
  query = "<nome do deputado> deputado <UF> instagram"
  procura o primeiro resultado do dominio instagram.com

Uso:
  python resolve_instagram.py --input deputados_estaduais.json
  python resolve_instagram.py --input deputados_federais.xlsx --formato xlsx

Dependencias:
  pip install requests beautifulsoup4 openpyxl lxml
"""

import argparse
import json
import re
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup

HEADERS = {"User-Agent": "Mozilla/5.0 Firefox/120.0"}
IG_RE = re.compile(r"instagram\.com/([A-Za-z0-9_.]+)")
BLACKLIST = {"explore", "p", "reel", "accounts", "about", "directory"}


def buscar_duckduckgo(query: str) -> str:
    """Retorna handle @xxx se achou, string vazia caso contrario."""
    url = "https://html.duckduckgo.com/html/"
    try:
        r = requests.post(url, data={"q": query}, headers=HEADERS, timeout=15)
        r.raise_for_status()
    except Exception as e:
        print(f"    busca falhou: {e}")
        return ""
    soup = BeautifulSoup(r.text, "lxml")
    for a in soup.select("a.result__url, a.result__a"):
        href = a.get("href") or ""
        m = IG_RE.search(href)
        if m and m.group(1).lower() not in BLACKLIST:
            return "@" + m.group(1)
    return ""


def enriquecer(items, campo_nome="nome", campo_uf="uf", delay=3.0):
    total = len(items)
    achados = 0
    for i, it in enumerate(items, 1):
        if it.get("instagram"):
            continue
        nome = it.get(campo_nome, "")
        uf = it.get(campo_uf, "")
        if not nome:
            continue
        q = f'"{nome}" deputad {uf} site:instagram.com'
        handle = buscar_duckduckgo(q)
        if handle:
            it["instagram"] = handle
            it["instagram_url"] = f"https://instagram.com/{handle.lstrip('@')}"
            achados += 1
        print(f"[{i}/{total}] {nome} ({uf}) -> {handle or 'nao achado'}")
        time.sleep(delay)
    print(f"\nResolvidos: {achados}/{total - sum(1 for i in items if not i.get(campo_nome))}")
    return items


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--formato", choices=["json", "xlsx"], default="json")
    ap.add_argument("--delay", type=float, default=3.0)
    args = ap.parse_args()

    if args.formato == "json":
        items = json.loads(Path(args.input).read_text(encoding="utf-8"))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(args.input)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            items.append(dict(zip(headers, row)))

    items = enriquecer(items, delay=args.delay)

    if args.formato == "json":
        Path(args.input).write_text(json.dumps(items, ensure_ascii=False, indent=2),
                                    encoding="utf-8")
    else:
        from openpyxl import load_workbook
        wb = load_workbook(args.input)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        # Adiciona coluna instagram se nao existir
        if "Instagram" not in headers:
            ws.cell(row=1, column=len(headers) + 1, value="Instagram")
        for i, it in enumerate(items, start=2):
            ws.cell(row=i, column=len(headers) + 1, value=it.get("instagram", ""))
        wb.save(args.input)

    print(f"\nArquivo atualizado: {args.input}")


if __name__ == "__main__":
    main()
