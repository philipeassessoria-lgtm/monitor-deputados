"""
coletar_posts.py
Coleta posts recentes (24h) de cada deputado usando Instaloader (biblioteca gratis).

Uso:
  python coletar_posts.py
  python coletar_posts.py --limite 10        # teste com so 10 perfis
  python coletar_posts.py --posts-por-perfil 3  # 3 posts recentes/perfil

Saida: posts_24h.json (e baixa thumbnails para ./thumbnails/)

Dependencias:
  pip install instaloader

ATENCAO:
  - Use conta IG DESCARTAVEL (nunca pessoal) via --login ou env IG_USER/IG_PASS.
  - Respeite delay (default 4s) para evitar rate-limit.
"""

import argparse
import json
import os
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import instaloader


def carregar_handles(path_federais="../deputados_federais.xlsx",
                     path_estaduais="../deputados_estaduais.json"):
    """Le os dois arquivos e retorna lista unificada [{nome, uf, handle}]."""
    handles = []
    # Estaduais (JSON)
    p_est = Path(path_estaduais)
    if p_est.exists():
        for d in json.loads(p_est.read_text(encoding="utf-8")):
            if d.get("instagram"):
                handles.append({
                    "nome": d["nome"], "uf": d.get("uf", ""),
                    "tipo": "estadual",
                    "handle": d["instagram"].lstrip("@"),
                })
    # Federais (XLSX)
    p_fed = Path(path_federais)
    if p_fed.exists():
        from openpyxl import load_workbook
        wb = load_workbook(p_fed)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        i_nome = headers.index("Nome") if "Nome" in headers else 1
        i_uf = headers.index("UF") if "UF" in headers else 4
        i_ig = headers.index("Instagram (handle)") if "Instagram (handle)" in headers else 6
        for row in ws.iter_rows(min_row=2, values_only=True):
            handle = (row[i_ig] or "").lstrip("@")
            if handle:
                handles.append({
                    "nome": row[i_nome], "uf": row[i_uf],
                    "tipo": "federal", "handle": handle,
                })
    return handles


def coletar(handles, posts_por_perfil=3, delay=4.0, janela_horas=24):
    L = instaloader.Instaloader(
        download_pictures=False, download_videos=False,
        download_video_thumbnails=False, download_geotags=False,
        download_comments=False, save_metadata=False,
        compress_json=False, post_metadata_txt_pattern="",
    )
    user = os.getenv("IG_USER")
    pw = os.getenv("IG_PASS")
    if user and pw:
        try:
            L.login(user, pw)
            print(f"Logado como {user}")
        except Exception as e:
            print(f"Login falhou: {e}. Seguindo sem login (rate-limit maior).")

    corte = datetime.now(timezone.utc) - timedelta(hours=janela_horas)
    resultados = []
    for i, h in enumerate(handles, 1):
        print(f"[{i}/{len(handles)}] @{h['handle']} ({h['nome']})")
        try:
            profile = instaloader.Profile.from_username(L.context, h["handle"])
            count = 0
            for post in profile.get_posts():
                if count >= posts_por_perfil:
                    break
                if post.date_utc < corte:
                    break
                engajamento = (post.likes or 0) + (post.comments or 0) * 3
                resultados.append({
                    "deputado": h["nome"],
                    "uf": h["uf"],
                    "tipo": h["tipo"],
                    "handle": h["handle"],
                    "post_id": post.shortcode,
                    "url": f"https://instagram.com/p/{post.shortcode}/",
                    "data_utc": post.date_utc.isoformat(),
                    "caption": (post.caption or "")[:500],
                    "likes": post.likes or 0,
                    "comments": post.comments or 0,
                    "engajamento": engajamento,
                    "is_video": post.is_video,
                    "thumbnail_url": post.url,
                    "mediacount_perfil": profile.mediacount,
                    "followers": profile.followers,
                })
                count += 1
            time.sleep(delay)
        except instaloader.exceptions.ProfileNotExistsException:
            print("   perfil nao existe")
        except instaloader.exceptions.PrivateProfileNotFollowedException:
            print("   perfil privado")
        except Exception as e:
            print(f"   erro: {e}")
            time.sleep(delay * 2)
    return resultados


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limite", type=int, help="Coleta so N perfis (teste)")
    ap.add_argument("--posts-por-perfil", type=int, default=3)
    ap.add_argument("--delay", type=float, default=4.0)
    ap.add_argument("--janela", type=int, default=24, help="Janela em horas")
    args = ap.parse_args()

    handles = carregar_handles()
    print(f"Total de handles carregados: {len(handles)}")
    if args.limite:
        handles = handles[:args.limite]

    posts = coletar(handles, posts_por_perfil=args.posts_por_perfil,
                    delay=args.delay, janela_horas=args.janela)
    Path("posts_24h.json").write_text(json.dumps(posts, ensure_ascii=False, indent=2),
                                      encoding="utf-8")
    print(f"\nSalvos {len(posts)} posts em posts_24h.json")


if __name__ == "__main__":
    main()
