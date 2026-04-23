"""
fetch_deputados_federais.py
Busca a lista completa dos 513 deputados federais do Brasil na legislatura atual,
extrai os perfis de Instagram (quando disponiveis na API) e gera um arquivo XLSX.

Fonte: API oficial da Camara dos Deputados
Docs:  https://dadosabertos.camara.leg.br/swagger/api.html

Requisitos:
    pip install requests openpyxl

Uso:
    python fetch_deputados_federais.py
Saida:
    deputados_federais.xlsx
"""

import re
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "https://dadosabertos.camara.leg.br/api/v2"
HEADERS = {"Accept": "application/json", "User-Agent": "deputados-ig-tracker/1.0"}

INSTAGRAM_RE = re.compile(
    r"(?:https?://)?(?:www\.)?instagram\.com/([A-Za-z0-9_.]+)/?",
    re.IGNORECASE,
)


def get_all_deputados():
    """Retorna a lista basica de todos os deputados em exercicio."""
    deputados = []
    url = f"{BASE}/deputados"
    params = {"itens": 100, "ordem": "ASC", "ordenarPor": "nome"}
    while url:
        r = requests.get(url, params=params, headers=HEADERS, timeout=30)
        r.raise_for_status()
        data = r.json()
        deputados.extend(data["dados"])
        next_link = next(
            (l["href"] for l in data.get("links", []) if l.get("rel") == "next"), None
        )
        url = next_link
        params = None  # parametros ja estao na URL de next
    return deputados


def get_detalhes(dep_id):
    """Retorna o objeto detalhado do deputado, incluindo redeSocial."""
    r = requests.get(f"{BASE}/deputados/{dep_id}", headers=HEADERS, timeout=30)
    r.raise_for_status()
    return r.json()["dados"]


def extrair_instagram(redes):
    """Recebe lista de strings de redes sociais e retorna (@handle, url)."""
    if not redes:
        return "", ""
    for item in redes:
        if not item:
            continue
        m = INSTAGRAM_RE.search(item)
        if m:
            handle = m.group(1).strip().rstrip("/")
            url = f"https://instagram.com/{handle}"
            return f"@{handle}", url
    return "", ""


def main():
    print("Buscando lista de deputados...")
    deputados = get_all_deputados()
    print(f"Encontrados {len(deputados)} deputados.")

    linhas = []
    for i, dep in enumerate(deputados, 1):
        try:
            det = get_detalhes(dep["id"])
            handle, url = extrair_instagram(det.get("redeSocial") or [])
            linhas.append({
                "id": dep["id"],
                "nome": det.get("ultimoStatus", {}).get("nomeEleitoral") or dep["nome"],
                "nome_civil": det.get("nomeCivil", ""),
                "partido": dep.get("siglaPartido", ""),
                "uf": dep.get("siglaUf", ""),
                "email": dep.get("email", ""),
                "instagram_handle": handle,
                "instagram_url": url,
                "site": next((s for s in det.get("redeSocial") or []
                              if s and "instagram" not in s.lower()
                              and "facebook" not in s.lower()
                              and "twitter" not in s.lower()
                              and "x.com" not in s.lower()), ""),
                "id_camara": dep["id"],
                "url_camara": dep.get("uri", ""),
            })
            print(f"[{i}/{len(deputados)}] {linhas[-1]['nome']} ({linhas[-1]['uf']}) -> {handle or 'sem IG'}")
            time.sleep(0.15)  # respeita o rate limit
        except Exception as e:
            print(f"ERRO em {dep.get('nome')}: {e}")

    # Gera XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Deputados Federais"

    headers = ["ID", "Nome", "Nome Civil", "Partido", "UF", "Email",
               "Instagram (handle)", "Instagram (URL)", "Site/Outro", "URL Camara"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="1F4E78")
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for linha in linhas:
        ws.append([
            linha["id"], linha["nome"], linha["nome_civil"], linha["partido"],
            linha["uf"], linha["email"], linha["instagram_handle"],
            linha["instagram_url"], linha["site"], linha["url_camara"],
        ])

    # Linha de totais
    total_row = len(linhas) + 3
    ws.cell(row=total_row, column=1, value="Total deputados:").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=f"=COUNTA(B2:B{len(linhas)+1})")
    ws.cell(row=total_row + 1, column=1, value="Com Instagram:").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=2, value=f'=COUNTIF(G2:G{len(linhas)+1},"<>""")')

    # Ajustes de coluna
    widths = [8, 30, 30, 10, 6, 35, 22, 40, 40, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, max_row=len(linhas) + 1):
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    wb.save("deputados_federais.xlsx")
    print(f"\nOK! {len(linhas)} deputados gravados em deputados_federais.xlsx")
    print(f"Com Instagram: {sum(1 for l in linhas if l['instagram_handle'])}")


if __name__ == "__main__":
    main()
